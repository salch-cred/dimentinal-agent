"""Document ingestion + table-aware retrieval.

Keeps coordinates (page / table / cell) for citation, and crucially preserves
*context*: the table title, the column header, and any footnote like
"$ in thousands" — that context is what lets the typed extractor assign the
correct `scale` to every figure.
"""
from __future__ import annotations

import os
import re
from typing import List

from schemas import Span


def _resolve_cell_val(ws, cell, visiting) -> float:
    coord = cell.coordinate
    if coord in visiting:
        return 0.0  # cycle detected
    visiting.add(coord)
    
    val = cell.value
    if val is None:
        visiting.remove(coord)
        return 0.0
    if isinstance(val, str) and val.startswith("="):
        res = _evaluate_formula(ws, val, visiting)
        visiting.remove(coord)
        return res
    
    visiting.remove(coord)
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def _expand_range(ws, range_str: str, visiting) -> float:
    total = 0.0
    try:
        cells = ws[range_str]
        if isinstance(cells, tuple):
            for row in cells:
                for cell in row:
                    total += _resolve_cell_val(ws, cell, visiting)
        else:
            total += _resolve_cell_val(ws, cells, visiting)
    except Exception:
        pass
    return total


def _evaluate_formula(ws, formula_str: str, visiting=None) -> float:
    if visiting is None:
        visiting = set()
    
    formula = formula_str.strip().upper()
    if formula.startswith("="):
        formula = formula[1:]
        
    # 1. Resolve SUM(A1:B3) or SUM(A1, B2)
    while True:
        m = re.search(r"SUM\(([^)]+)\)", formula)
        if not m:
            break
        sum_arg_str = m.group(1)
        args = sum_arg_str.split(",")
        sum_val = 0.0
        for arg in args:
            arg = arg.strip()
            if ":" in arg:
                sum_val += _expand_range(ws, arg, visiting)
            else:
                if re.match(r"^[A-Z]+\d+$", arg):
                    cell = ws[arg]
                    sum_val += _resolve_cell_val(ws, cell, visiting)
                else:
                    try:
                        sum_val += float(arg)
                    except ValueError:
                        sum_val += _evaluate_formula(ws, "=" + arg, visiting)
        formula = formula[:m.start()] + str(sum_val) + formula[m.end():]
        
    # 2. Resolve single cell references
    while True:
        m = re.search(r"\b([A-Z]+\d+)\b", formula)
        if not m:
            break
        coord = m.group(1)
        cell = ws[coord]
        val = _resolve_cell_val(ws, cell, visiting)
        formula = formula[:m.start()] + str(val) + formula[m.end():]
        
    # 3. Safe arithmetic evaluation
    if not re.match(r"^[0-9.+\-*/()\s]+$", formula):
        return 0.0
        
    try:
        return float(eval(formula, {"__builtins__": {}}))
    except Exception:
        return 0.0

# Tokens that signal a scale qualifier; the extractor reads these from context.
_SCALE_HINTS = {
    "thousands": 1e3,
    "million": 1e6,
    "millions": 1e6,
    "billion": 1e9,
    "billions": 1e9,
}


def _detect_scale(context: str) -> float:
    """Best-effort: read 'in thousands' / 'in millions' style hints."""
    low = context.lower()
    for word, scale in _SCALE_HINTS.items():
        if re.search(rf"\b(in\s+)?{word}\b", low):
            return scale
    return 1.0


def _clean(text: str) -> str:
    return re.sub(r"\s+", " ", (text or "")).strip()


_PERIOD_TOKEN = re.compile(r"\b(FY\s*\d{4}|Q[1-4][\s\-]?\d{4}|as_of_\d|FY\d{4}|Q[1-4]\b|\b(19|20)\d{2}\b)", re.I)


def _looks_like_header(cells: List[str]) -> bool:
    """A header row has period-like tokens (FY2024, Q4-2024, as_of_...) in >=2 cells."""
    hits = sum(1 for c in cells if c and _PERIOD_TOKEN.search(c))
    return hits >= 2


def _sheet_scale_hint(ws_title: str, all_text: str) -> str:
    """Find a '$ in thousands/millions' style footnote anywhere on the sheet."""
    low = all_text.lower()
    for word in ("thousands", "millions", "billions"):
        if re.search(rf"\bin\s+{word}\b", low):
            return f"$ in {word}"
    return ""


def _make_markdown_table_spans(ws, rows, hint: str) -> List[Span]:
    if not rows:
        return []
    import openpyxl
    lines = []
    max_cols = max(len(r) for r in rows) if rows else 0
    if max_cols == 0:
        return []
    for ri, row in enumerate(rows):
        row_str_parts = []
        for ci in range(max_cols):
            val = row[ci] if ci < len(row) else None
            cell_ref = f"{openpyxl.utils.get_column_letter(ci+1)}{ri+1}"
            if val is None or not str(val).strip():
                row_str_parts.append("")
            else:
                display = str(val)
                try:
                    cell_obj = ws.cell(row=ri + 1, column=ci + 1)
                    nf = (cell_obj.number_format or "")
                    if "%" in nf:
                        display = f"{float(val) * 100:g}%"
                except Exception:
                    pass
                row_str_parts.append(f"{_clean(display)} [{cell_ref}]")
        lines.append("| " + " | ".join(row_str_parts) + " |")
        if ri == 0:
            lines.append("| " + " | ".join(["---"] * max_cols) + " |")
    spans = []
    chunk_size = 50
    for i in range(0, len(lines), chunk_size):
        chunk_lines = lines[i:i+chunk_size]
        if i > 0 and len(lines) > 2:
            chunk_lines = [lines[0], lines[1]] + chunk_lines
        chunk_text = "\n".join(chunk_lines)
        spans.append(Span(
            id=f"{ws.title}__layout_chunk_{i//chunk_size}",
            text=chunk_text,
            context=f"sheet={ws.title} layout=table chunk={i//chunk_size} scale={hint or '1.0'}",
            source=f"{ws.title}!layout_chunk_{i//chunk_size}",
        ))
    return spans


def _make_csv_markdown_table_spans(path: str, reader: List[List[str]]) -> List[Span]:
    if not reader:
        return []
    lines = []
    max_cols = max(len(r) for r in reader) if reader else 0
    if max_cols == 0:
        return []
    for ri, row in enumerate(reader):
        row_str_parts = []
        for ci in range(max_cols):
            val = row[ci] if ci < len(row) else ""
            cell_ref = f"r{ri+1}c{ci+1}"
            if not val.strip():
                row_str_parts.append("")
            else:
                row_str_parts.append(f"{_clean(val)} [{cell_ref}]")
        lines.append("| " + " | ".join(row_str_parts) + " |")
        if ri == 0:
            lines.append("| " + " | ".join(["---"] * max_cols) + " |")
    spans = []
    chunk_size = 50
    base_name = os.path.basename(path)
    for i in range(0, len(lines), chunk_size):
        chunk_lines = lines[i:i+chunk_size]
        if i > 0 and len(lines) > 2:
            chunk_lines = [lines[0], lines[1]] + chunk_lines
        chunk_text = "\n".join(chunk_lines)
        spans.append(Span(
            id=f"csv__layout_chunk_{i//chunk_size}",
            text=chunk_text,
            context=f"file={base_name} layout=table chunk={i//chunk_size}",
            source=f"file={base_name} layout_chunk_{i//chunk_size}",
        ))
    return spans


def _make_pdf_markdown_table_spans(page_no: int, ti: int, table: List[List[str]]) -> List[Span]:
    if not table:
        return []
    lines = []
    max_cols = max(len(r) for r in table) if table else 0
    if max_cols == 0:
        return []
    for ri, row in enumerate(table):
        row_str_parts = []
        for ci in range(max_cols):
            val = row[ci] if ci < len(row) else ""
            cell_ref = f"r{ri+1}c{ci+1}"
            if val is None or not str(val).strip():
                row_str_parts.append("")
            else:
                row_str_parts.append(f"{_clean(str(val))} [{cell_ref}]")
        lines.append("| " + " | ".join(row_str_parts) + " |")
        if ri == 0:
            lines.append("| " + " | ".join(["---"] * max_cols) + " |")
    chunk_text = "\n".join(lines)
    return [Span(
        id=f"pdf__p{page_no}_t{ti}_layout",
        text=chunk_text,
        context=f"page={page_no} table={ti} layout=table",
        source=f"page={page_no} table={ti} layout",
    )]


def _get_merged_cell_value(ws, cell) -> Any:
    val = cell.value
    if val is not None:
        return val
    # Check if this cell is within any merged ranges
    for r in ws.merged_cells.ranges:
        if cell.coordinate in r:
            # Top-left cell of the merged range
            return ws.cell(row=r.min_row, column=r.min_col).value
    return None


def _local_scale_hint(rows: List[List[Any]], current_row_idx: int, sheet_hint: str) -> str:
    """Scan 5 rows above/below current_row_idx for local scale overrides."""
    start = max(0, current_row_idx - 5)
    end = min(len(rows), current_row_idx + 6)
    for ri in range(start, end):
        row_text = " ".join(str(v) for v in (rows[ri] or []) if v is not None).lower()
        for word in ("thousands", "millions", "billions"):
            if re.search(rf"\bin\s+{word}\b", row_text):
                return f"$ in {word}"
    return sheet_hint


def ingest_xlsx(path: str) -> List[Span]:
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=False)
    spans: List[Span] = []
    for ws in wb.worksheets:
        rows_cells = list(ws.iter_rows(values_only=False))
        rows = []
        for row_c in rows_cells:
            row_vals = []
            for cell in row_c:
                val = _get_merged_cell_value(ws, cell)
                if isinstance(val, str) and val.startswith("="):
                    val = _evaluate_formula(ws, val)
                row_vals.append(val)
            rows.append(row_vals)

        # Pre-scan all text on the sheet for a scale footnote.
        sheet_text = " ".join(
            str(v) for row in rows for v in (row or []) if v is not None
        )
        sheet_hint = _sheet_scale_hint(ws.title, sheet_text)

        # Find the header row: first row that looks like a period header.
        header_idx = None
        header: List[str] = []
        for ri, row in enumerate(rows):
            cells = [("" if v is None else str(v)) for v in (row or [])]
            if _looks_like_header(cells):
                header_idx = ri
                header = cells
                break
        if header_idx is None:
            # No header detected — fall back: treat each row generically.
            for ri, row in enumerate(rows):
                for ci, v in enumerate(row or []):
                    if v is None or not str(v).strip():
                        continue
                    spans.append(Span(
                        id=f"{ws.title}__r{ri}c{ci}",
                        text=_clean(str(v)),
                        context=f"sheet={ws.title}",
                        source=f"{ws.title}!{openpyxl.utils.get_column_letter(ci+1)}{ri+1}",
                    ))
            # Generate layout chunks for this ws
            spans.extend(_make_markdown_table_spans(ws, rows, sheet_hint))
            continue

        # Data rows come after the header. Column A is the metric label.
        current_parent = ""
        for ri in range(header_idx + 1, len(rows)):
            raw_row = rows[ri] or []
            row = [("" if v is None else str(v)) for v in raw_row]
            if not any(c.strip() for c in row):
                current_parent = ""  # Reset parent category on empty separator rows
                continue
            metric_label = _clean(row[0]) if row else ""
            
            # Category header heuristic: Column A is populated, but other columns are empty.
            if metric_label and not any(c.strip() for c in row[1:]):
                current_parent = metric_label
                continue
                
            full_metric_label = metric_label
            if current_parent and metric_label and metric_label != current_parent:
                full_metric_label = f"{current_parent} — {metric_label}"
                
            # Scan for a local scale hint near this row
            row_hint = _local_scale_hint(rows, ri, sheet_hint)

            for ci, val in enumerate(row):
                if not val.strip():
                    continue
                col_name = _clean(header[ci]) if ci < len(header) else ""
                # Detect percent-formatted cells (e.g. number_format "0.00%").
                # A raw 0.4087 in such a cell displays as 40.87%; render it as
                # a percentage so the value carries its true magnitude.
                display = val
                is_percent = False
                if 0 < ci and ci < len(raw_row):
                    cell_obj = ws.cell(row=ri + 1, column=ci + 1)
                    nf = (cell_obj.number_format or "")
                    if "%" in nf:
                        try:
                            display = f"{float(val) * 100:g}"
                            is_percent = True
                        except ValueError:
                            pass
                
                # Heuristic: if the row contains "margin"/"rate"/"%" and the value
                # is a float between -1.0 and 1.0 (e.g. 0.152), scale it to percent (15.2%)
                if 0 < ci and ci < len(raw_row) and not is_percent:
                    metric_lower = full_metric_label.lower()
                    col_lower = col_name.lower()
                    is_margin_term = any(term in metric_lower or term in col_lower for term in ("margin", "rate", "%", "percentage"))
                    if is_margin_term:
                        try:
                            fval = float(val)
                            if -1.0 <= fval <= 1.0 and fval != 0.0:
                                display = f"{fval * 100:g}"
                                is_percent = True
                        except ValueError:
                            pass

                parts = [f"sheet={ws.title}"]
                if full_metric_label:
                    parts.append(f"metric={full_metric_label}")
                if col_name:
                    parts.append(f"period={col_name}")
                if row_hint and not is_percent:
                    parts.append(row_hint)  # "$ in thousands" doesn't apply to a %
                if is_percent:
                    parts.append("unit=percent")
                ctx = " ".join(parts)
                cell_ref = f"{ws.title}!{openpyxl.utils.get_column_letter(ci+1)}{ri+1}"
                # Rich text: include the metric + period so retrieval & the
                # extractor both see the figure in full context.
                rich = display
                if full_metric_label and ci > 0:
                    rich = f"{full_metric_label} ({col_name}): {display}"
                spans.append(Span(
                    id=f"{ws.title}__{openpyxl.utils.get_column_letter(ci+1)}{ri+1}",
                    text=_clean(rich),
                    context=ctx,
                    source=cell_ref,
                ))
        # Generate layout chunks for this ws
        spans.extend(_make_markdown_table_spans(ws, rows, sheet_hint))
    return spans


def ingest_csv(path: str) -> List[Span]:
    import csv

    spans: List[Span] = []
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = list(csv.reader(f))
    if not reader:
        return spans
    header = reader[0]
    for ri, row in enumerate(reader[1:], start=2):
        for ci, val in enumerate(row):
            if not val.strip():
                continue
            col_name = header[ci] if ci < len(header) else f"col{ci}"
            spans.append(Span(
                id=f"csv__r{ri}c{ci}",
                text=_clean(val),
                context=f"column={_clean(col_name)}",
                source=f"row={ri} col={_clean(col_name)}",
            ))
    # Generate layout chunks for this CSV
    spans.extend(_make_csv_markdown_table_spans(path, reader))
    return spans


def ingest_pdf(path: str) -> List[Span]:
    """Extract tables cell-by-cell with pdfplumber, keeping page + table coords."""
    import pdfplumber

    spans: List[Span] = []
    with pdfplumber.open(path) as pdf:
        for page_no, page in enumerate(pdf.pages, start=1):
            tables = page.extract_tables() or []
            for ti, table in enumerate(tables, start=1):
                header: List[str] = []
                # Footnote / scale hints from the page text (cheap heuristic).
                page_text = page.extract_text() or ""
                scale_hint = _detect_scale(page_text)
                hint_ctx = ""
                for word, sc in _SCALE_HINTS.items():
                    if sc == scale_hint and re.search(rf"\bin\s+{word}\b", page_text.lower()):
                        hint_ctx = f"$ in {word}"
                        break

                for ri, row in enumerate(table):
                    cells = [("" if c is None else str(c)) for c in (row or [])]
                    if ri == 0 or not header:
                        if any(c.strip() for c in cells):
                            header = cells
                            continue
                    for ci, val in enumerate(cells):
                        if not val.strip():
                            continue
                        col_name = header[ci] if ci < len(header) else f"col{ci}"
                        ctx_parts = [f"page={page_no}", f"col={_clean(col_name)}"]
                        if hint_ctx:
                            ctx_parts.append(hint_ctx)
                        spans.append(Span(
                            id=f"p{page_no}_t{ti}_r{ri}_c{ci}",
                            text=_clean(val),
                            context=" ".join(ctx_parts),
                            source=f"page={page_no} table={ti} row={ri} col={_clean(col_name)}",
                        ))
                # Generate layout chunks for this PDF table
                spans.extend(_make_pdf_markdown_table_spans(page_no, ti, table))
            # Also keep plain text lines as spans (covers prose like footnotes).
            text = page.extract_text() or ""
            for li, line in enumerate(text.splitlines(), start=1):
                line = _clean(line)
                if not line:
                    continue
                # Avoid duplicating pure-table numbers already captured above;
                # keep lines that look like sentences/notes.
                if len(line.split()) >= 4 and not re.fullmatch(r"[\d,$\s.\-()%]+", line):
                    spans.append(Span(
                        id=f"p{page_no}_l{li}",
                        text=line,
                        context=f"page={page_no} text",
                        source=f"page={page_no} line={li}",
                    ))
    return spans


# Allowed document roots (resolved at import). A doc_path must resolve to a
# real path inside one of these, after symlink/`..` normalization, or it is
# rejected. Override via the DOCS_ROOTS env var (os.pathsep-separated).
_DEFAULT_ROOTS = [
    os.path.join(os.path.dirname(os.path.abspath(__file__)), "sample_data"),
    "/data",
    "./data",
    "/workspace/data",
    "/workspace",
    "./corpus",
    "/corpus",
]


DOCS_ROOTS = []
for _r in os.environ.get("DOCS_ROOTS", "").split(os.pathsep):
    _r = _r.strip()
    if _r:
        DOCS_ROOTS.append(os.path.abspath(_r))
if not DOCS_ROOTS:
    DOCS_ROOTS = _DEFAULT_ROOTS

_ALLOWED_EXTS = {".xlsx", ".xlsm", ".csv", ".pdf", ".txt", ".json"}


class UnsafePathError(ValueError):
    """Raised when doc_path escapes the allowed roots or extensions."""


def is_safe_path(path: str) -> bool:
    """True if `path` resolves inside an allowed root AND has an allowed ext."""
    if not path or not isinstance(path, str):
        return False
    ext = os.path.splitext(path)[1].lower()
    if ext not in _ALLOWED_EXTS:
        return False
    real = os.path.realpath(path)
    for root in DOCS_ROOTS:
        root_real = os.path.realpath(root)
        try:
            common = os.path.commonpath([root_real, real])
        except ValueError:
            continue
        if common == root_real:
            return True
    return False


def resolve_doc_path(path: str) -> str:
    """Validate + canonicalize a user-supplied doc_path, or raise."""
    if not is_safe_path(path):
        raise UnsafePathError(
            f"doc_path not allowed (must be under an allowed root and end in "
            f"{sorted(_ALLOWED_EXTS)}): {path!r}"
        )
    real = os.path.realpath(path)
    if not os.path.exists(real):
        raise FileNotFoundError(path)
    return real


def ingest_txt(path: str) -> List[Span]:
    spans: List[Span] = []
    with open(path, "r", encoding="utf-8") as f:
        for li, line in enumerate(f, start=1):
            line = _clean(line)
            if not line:
                continue
            spans.append(Span(
                id=f"txt__l{li}",
                text=line,
                context=f"file={os.path.basename(path)}",
                source=f"file={os.path.basename(path)} line={li}",
            ))
    return spans


def ingest_json(path: str) -> List[Span]:
    import json
    spans: List[Span] = []
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        
        def extract_strings(obj, context=""):
            if isinstance(obj, str):
                val = _clean(obj)
                if val:
                    spans.append(Span(
                        id=f"json__s{len(spans)}",
                        text=val,
                        context=context.strip(),
                        source=f"file={os.path.basename(path)}"
                    ))
            elif isinstance(obj, list):
                for i, item in enumerate(obj):
                    extract_strings(item, f"{context} list[{i}]")
            elif isinstance(obj, dict):
                for k, v in obj.items():
                    extract_strings(v, f"{context} {k}")
        extract_strings(data, f"file={os.path.basename(path)}")
    except Exception:
        pass
    return spans


def ingest(path: str) -> List[Span]:
    """Dispatch on extension. Raises ValueError on unsupported types."""
    path = resolve_doc_path(path)  # confine to allowed roots/extensions
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        return ingest_xlsx(path)
    if ext == ".csv":
        return ingest_csv(path)
    if ext == ".pdf":
        return ingest_pdf(path)
    if ext == ".txt":
        return ingest_txt(path)
    if ext == ".json":
        return ingest_json(path)
    raise ValueError(f"unsupported file type: {ext}")


# --------------------------------------------------------------------------- #
#  Retrieval                                                                   #
# --------------------------------------------------------------------------- #
_STOP = set("a an the of for to in on at and or is are was were be been being "
            "this that these those it its as by with from per vs versus what "
            "how many much did does do company's company".split())

_SYNONYMS = {
    "revenue": {"sales", "turnover", "inflow"},
    "sales": {"revenue", "turnover", "inflow"},
    "turnover": {"revenue", "sales", "inflow"},
    "income": {"earnings", "profit", "net"},
    "earnings": {"income", "profit", "net"},
    "profit": {"income", "earnings", "net"},
    "cash": {"liquidity", "equivalents", "cash-flow"},
    "growth": {"change", "increase", "decrease", "yoy", "growth-rate"},
    "margin": {"ratio", "percent", "percentage", "profitability"},
    "debt": {"borrowing", "loan", "liabilities", "obligation"},
    "equity": {"stockholders", "shareholders", "capital"},
    "research": {"development", "rd", "r&d"},
    "development": {"research", "rd", "r&d"},
    "rd": {"research", "development", "r&d"},
    "spending": {"expenditure", "expense", "expenses", "cost", "outlay"},
    "expenditure": {"spending", "expense", "expenses", "cost", "outlay"},
    "expenses": {"spending", "expenditure", "expense", "cost", "outlay"},
    "expense": {"spending", "expenditure", "expenses", "cost", "outlay"},
    "assets": {"property", "goodwill", "investments", "resources"},
    "operating": {"operations", "opex"},
    "free": {"fcf", "cashflow"},
    "dividends": {"dividend", "payout", "distribution"},
}


def _tokens(text: str) -> set:
    text_clean = text.lower().replace("r&d", "rd").replace("r & d", "rd")
    return {w for w in re.findall(r"[a-z0-9]+", text_clean) if w not in _STOP and len(w) > 1}


# Key financial metrics that should be strongly preferred over noise rows
_KEY_METRICS = {
    "total revenue", "cost of revenue", "gross profit", "operating income",
    "net income", "operating expenses", "interest expense", "cash flow",
    "free cash flow", "operating cash flow", "capital expenditures",
    "total assets", "total liabilities", "stockholders' equity",
    "current assets", "current liabilities", "long-term debt",
    "cash and cash equivalents", "diluted eps", "dividends",
    "research", "development", "goodwill", "accounts receivable",
    "gross margin", "operating margin", "segment", "cloud", "enterprise",
    "americas", "europe", "asia", "emea", "apac",
}


def retrieve(question: str, spans: List[Span], k: int = 15) -> List[Span]:
    """Rank spans by keyword overlap with the question; return top k.

    Boosts spans that contain a digit (financial QA is almost always about
    numbers) and spans whose scale-hint context matches a magnitude word in
    the question (e.g. "in millions").
    """
    qtoks = _tokens(question)
    if not qtoks:
        return spans[:k]

    # Increase k for ratio/percentage/cross-sheet queries
    q_lower = question.lower()
    needs_more = any(w in q_lower for w in
                     ("ratio", "percentage", "% of", "as a percentage",
                      "margin", "debt-to", "plus", "combined"))
    effective_k = max(k, 25) if needs_more else k

    # Expand query tokens with common financial synonyms
    expanded_qtoks = set(qtoks)
    for tok in qtoks:
        if tok in _SYNONYMS:
            expanded_qtoks.update(_SYNONYMS[tok])

    scored = []
    for sp in spans:
        stoks = _tokens(sp.text) | _tokens(sp.context)
        overlap = len(expanded_qtoks & stoks)
        score = overlap
        if re.search(r"\d", sp.text):
            score += 0.5
        # Magnitude-word boost: if the question mentions a magnitude and the
        # span's context carries the matching scale hint, prefer it.
        for word in _SCALE_HINTS:
            if word in question.lower() and word in sp.context.lower():
                score += 1.0
        # Boost key financial metrics over noise rows like "Expense Line Item 1234"
        sp_lower = sp.text.lower()
        if any(m in sp_lower for m in _KEY_METRICS):
            score += 2.0
        # Penalize noise rows (generic expense/asset line items)
        if re.search(r"(expense line item|asset line)\s*\d+", sp_lower):
            score -= 5.0
        # Boost layout chunk spans if the sheet/file matches the query
        if "layout=table" in sp.context:
            sheet_title_match = False
            for word in qtoks:
                if word in sp.context.lower() or word in sp_lower:
                    sheet_title_match = True
                    break
            if sheet_title_match:
                score += 3.0
        if score > 0:
            scored.append((score, sp))

    scored.sort(key=lambda t: t[0], reverse=True)
    return [sp for _, sp in scored[:effective_k]] if scored else spans[:effective_k]
