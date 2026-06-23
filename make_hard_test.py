import os
import openpyxl
from openpyxl.styles import Font

def generate_hard_sheet():
    # Load with data_only=False so we can write formulas and evaluate them locally
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Global Operations"
    
    # Title & Metadata
    ws["A1"] = "ACME Global Operations Statement"
    ws["A2"] = "(Amounts in thousands)"
    ws["A1"].font = Font(bold=True, size=13)
    
    # Headers
    headers = ["Metric Name", "FY2023", "FY2024"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=4, column=c, value=h).font = Font(bold=True)
        
    # We will build a nested structure:
    # Revenue
    #   US Division (USD)
    #   EU Division (EUR)
    #   Total Revenue (Formula)
    #
    # Costs
    #   US Cost (USD)
    #   EU Cost (EUR)
    #   Total Cost (Formula)
    #
    # Net Profit (Formula)
    
    ws["A5"] = "Revenue"
    ws["A5"].font = Font(bold=True)
    
    ws["A6"] = "US Division"
    ws["B6"] = 1000000 # 1,000,000 thousands = 1B USD
    ws["C6"] = 1200000 # 1.2B USD
    
    ws["A7"] = "EU Division"
    ws["B7"] = 800000 # 800M EUR
    ws["C7"] = 900000 # 900M EUR
    # Mark the currency context on the cell if possible, or we will query it.
    
    ws["A8"] = "Total Revenue"
    ws["B8"] = "=B6+B7" # Formula
    ws["C8"] = "=C6+C7" # Formula
    
    ws["A10"] = "Costs"
    ws["A10"].font = Font(bold=True)
    
    ws["A11"] = "US Cost"
    ws["B11"] = 600000
    ws["C11"] = 700000
    
    ws["A12"] = "EU Cost"
    ws["B12"] = 500000
    ws["C12"] = 550000
    
    ws["A13"] = "Total Cost"
    ws["B13"] = "=B11+B12"
    ws["C13"] = "=C11+C12"
    
    ws["A15"] = "Net Profit"
    ws["A15"].font = Font(bold=True)
    ws["B15"] = "=B8-B13"
    ws["C15"] = "=C8-C13"
    
    out_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "sample_data")
    os.makedirs(out_dir, exist_ok=True)
    path = os.path.join(out_dir, "acme_global_hard.xlsx")
    wb.save(path)
    print(f"Generated advanced hard spreadsheet at: {path}")

if __name__ == "__main__":
    generate_hard_sheet()
