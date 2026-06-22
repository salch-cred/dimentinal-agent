"""Generate the sample financial document used for the smoke test and demo.

Creates sample_data/acme_financials.xlsx — a deliberately trap-laden sheet:
  * A "$ in thousands" footnote (scale trap)
  * Q4 column next to FY columns (period trap)
  * Cloud Segment rows next to Consolidated rows (entity trap)
  * Balance-sheet items (cash) next to income-statement flows (flow/stock trap)

Run:  python make_sample_data.py
"""
from __future__ import annotations

import os

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


def build() -> str:
    wb = openpyxl.Workbook()

    # ---- Income Statement (in thousands) ----
    ws = wb.active
    ws.title = "Income Statement"
    ws["A1"] = "ACME Corporation — Consolidated Statements of Operations"
    ws["A2"] = "(Amounts in thousands of U.S. dollars, except per-share data)"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"].font = Font(italic=True, color="555555")

    headers = ["", "FY2023", "FY2024", "Q4-2024"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    rows = [
        ("Total revenue",              4_510_000, 4_820_000, 1_240_000),
        ("Cost of revenue",            2_700_000, 2_850_000,   740_000),
        ("Gross profit",               1_810_000, 1_970_000,   500_000),
        ("Operating expenses",         1_200_000, 1_310_000,   345_000),
        ("Operating income",             610_000,   660_000,   155_000),
        ("Net income (consolidated)",    480_000,   520_000,   140_000),
    ]
    # NOTE the trap: a "Cloud Segment" net-income row sits right below consolidated.
    segment_rows = [
        ("  Cloud Segment — net income", 190_000,   215_000,    58_000),
    ]
    r = 5
    for label, *vals in rows + segment_rows:
        ws.cell(row=r, column=1, value=label)
        for c, v in enumerate(vals, start=2):
            cell = ws.cell(row=r, column=c, value=v)
            cell.number_format = "#,##0"
        r += 1
    ws.cell(row=r, column=1, value="Gross margin %").font = Font(italic=True)
    ws.cell(row=r, column=2, value=0.4012).number_format = "0.00%"
    ws.cell(row=r, column=3, value=0.4087).number_format = "0.00%"
    ws.cell(row=r, column=4, value=0.4032).number_format = "0.00%"

    # ---- Balance Sheet (point-in-time) ----
    ws2 = wb.create_sheet("Balance Sheet")
    ws2["A1"] = "ACME Corporation — Consolidated Balance Sheet"
    ws2["A2"] = "(Amounts in thousands of U.S. dollars)"
    ws2["A1"].font = Font(bold=True, size=13)
    ws2["A2"].font = Font(italic=True, color="555555")
    for c, h in enumerate(["", "as_of_2023-12-31", "as_of_2024-12-31"], start=1):
        cell = ws2.cell(row=4, column=c, value=h)
        cell.font = Font(bold=True); cell.alignment = Alignment(horizontal="center")
    bs_rows = [
        ("Cash and cash equivalents", 1_120_000, 1_350_000),
        ("Total current assets",      3_200_000, 3_650_000),
        ("Total assets",              8_900_000, 9_400_000),
        ("Total stockholders' equity",5_100_000, 5_450_000),
    ]
    r = 5
    for label, v23, v24 in bs_rows:
        ws2.cell(row=r, column=1, value=label)
        ws2.cell(row=r, column=2, value=v23).number_format = "#,##0"
        ws2.cell(row=r, column=3, value=v24).number_format = "#,##0"
        r += 1

    # ---- Per-share ----
    ws3 = wb.create_sheet("Per Share")
    ws3["A1"] = "ACME Corporation — Per Share Data"
    ws3["A1"].font = Font(bold=True, size=13)
    for c, h in enumerate(["", "FY2023", "FY2024"], start=1):
        cell = ws3.cell(row=3, column=c, value=h)
        cell.font = Font(bold=True); cell.alignment = Alignment(horizontal="center")
    ws3.cell(row=4, column=1, value="Diluted shares outstanding (millions)")
    ws3.cell(row=4, column=2, value=420); ws3.cell(row=4, column=3, value=418)
    ws3.cell(row=5, column=1, value="Diluted EPS")
    ws3.cell(row=5, column=2, value=1.14); ws3.cell(row=5, column=3, value=1.24)

    # widen columns
    for sheet in wb.worksheets:
        sheet.column_dimensions["A"].width = 38
        for col in ["B", "C", "D"]:
            sheet.column_dimensions[col].width = 18

    out_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "sample_data")
    os.makedirs(out_dir, exist_ok=True)
    out = os.path.join(out_dir, "acme_financials.xlsx")
    wb.save(out)
    return out


if __name__ == "__main__":
    path = build()
    print(f"Wrote {path}")
    # Sanity: re-open and print.
    wb = openpyxl.load_workbook(path, data_only=True)
    for ws in wb.worksheets:
        print(f"\n=== {ws.title} ===")
        for row in ws.iter_rows(values_only=True):
            if any(v is not None for v in row):
                print("  ", row)
