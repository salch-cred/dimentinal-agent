import os
import time
import openpyxl
from openpyxl.styles import Font
from ingest import ingest
from app import run_proof_pipeline

def generate_large_sheet():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Large Financials"
    
    # Scale footnote
    ws["A1"] = "ACME Corporation — Large Data Stress Test"
    ws["A2"] = "(Amounts in thousands of U.S. dollars)"
    ws["A1"].font = Font(bold=True, size=13)
    
    # Headers
    headers = ["Metric Name", "FY2023", "FY2024"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=4, column=c, value=h).font = Font(bold=True)
        
    # Generate 1000 rows of financial figures
    # We will include regular names and a few unique ones at the end to search for
    for r in range(5, 1005):
        ws.cell(row=r, column=1, value=f"Financial Expense Category {r-4}")
        ws.cell(row=r, column=2, value=100 + r)
        ws.cell(row=r, column=3, value=150 + r)
        
    # Add a target metric we want to query at the very end
    ws.cell(row=1005, column=1, value="Consolidated R&D Expenditure")
    ws.cell(row=1005, column=2, value=45000) # In thousands
    ws.cell(row=1005, column=3, value=52000) # In thousands
    
    # Make directory if not exists
    out_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "sample_data")
    os.makedirs(out_dir, exist_ok=True)
    
    path = os.path.join(out_dir, "acme_financials_1000.xlsx")
    wb.save(path)
    return path

def run_performance_test():
    print("=== 1. Generating 1000-Row Spreadsheet ===")
    t0 = time.time()
    path = generate_large_sheet()
    t_gen = time.time() - t0
    print(f"Spreadsheet generated successfully in {t_gen:.3f}s: {path}\n")
    
    print("=== 2. Benchmarking Ingestion Pipeline ===")
    t0 = time.time()
    spans = ingest(path)
    t_ingest = time.time() - t0
    print(f"Successfully ingested {len(spans)} cell spans in {t_ingest:.3f}s")
    print(f"Average ingestion speed: {len(spans)/t_ingest:.1f} spans/second\n")
    
    print("=== 3. Benchmarking Search & Proof Resolution ===")
    question = "What was the Consolidated R&D Expenditure for FY2024?"
    print(f"Query: \"{question}\"")
    
    t0 = time.time()
    proof = run_proof_pipeline(question, path)
    t_pipeline = time.time() - t0
    
    print("\n--- Verified Output Receipt ---")
    print(f"Status: {'Rejected' if proof.rejected else 'Verified'}")
    print(f"Calculated Answer: {proof.answer}")
    print(f"Citations: {proof.citations}")
    print(f"Dimension Checks: {proof.dimension_checks}")
    print(f"Verifier Verdict: {proof.verifier_verdict}")
    print(f"Total Pipeline Execution Time: {t_pipeline:.3f}s")
    
    # Clean up the generated file to keep workspace clean (optional but good practice)
    # If the user wants to test it on the UI, they can leave the file there.
    # So we will KEEP the file under sample_data so they can upload it on the frontend!
    print(f"\nSaved file at '{path}' so you can upload and test it in your browser UI!")

if __name__ == "__main__":
    run_performance_test()
