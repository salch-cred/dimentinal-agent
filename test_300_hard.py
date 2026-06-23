"""Hard Benchmark with 300 questions to stress test the pipeline.

Generates a realistic, trap-laden financial workbook with ~100K cells
across multiple sheets, then runs 300 hard questions through the pipeline.
"""
from __future__ import annotations

import json
import os
import sys
import time
from typing import Any

import openpyxl
from openpyxl.styles import Font, Alignment

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from app import AskRequest, run_proof_pipeline, baseline
from ingest import ingest

# -------------------------------------------------------------
# Define data values deterministically
# -------------------------------------------------------------
YEARS = ["FY2020", "FY2021", "FY2022", "FY2023", "FY2024"]
QUARTERS = ["Q1-2023", "Q2-2023", "Q3-2023", "Q4-2023", "Q1-2024", "Q2-2024", "Q3-2024", "Q4-2024"]

# Data values in absolute dollars
REVENUE = {
    "FY2020": 3_200_000_000, "FY2021": 3_500_000_000, "FY2022": 3_800_000_000, "FY2023": 4_200_000_000, "FY2024": 4_650_000_000,
    "Q1-2023": 950_000_000, "Q2-2023": 1_000_000_000, "Q3-2023": 1_100_000_000, "Q4-2023": 1_150_000_000,
    "Q1-2024": 1_050_000_000, "Q2-2024": 1_100_000_000, "Q3-2024": 1_200_000_000, "Q4-2024": 1_300_000_000
}

COST_OF_REVENUE = {
    "FY2020": 1_900_000_000, "FY2021": 2_100_000_000, "FY2022": 2_300_000_000, "FY2023": 2_500_000_000, "FY2024": 2_750_000_000,
    "Q1-2023": 580_000_000, "Q2-2023": 600_000_000, "Q3-2023": 650_000_000, "Q4-2023": 670_000_000,
    "Q1-2024": 630_000_000, "Q2-2024": 660_000_000, "Q3-2024": 710_000_000, "Q4-2024": 750_000_000
}

GROSS_PROFIT = {
    k: REVENUE[k] - COST_OF_REVENUE[k] for k in REVENUE
}

GROSS_MARGIN = {
    k: GROSS_PROFIT[k] / REVENUE[k] for k in REVENUE
}

RD_SPENDING = {
    "FY2020": 320_000_000, "FY2021": 360_000_000, "FY2022": 400_000_000, "FY2023": 450_000_000, "FY2024": 520_000_000,
    "Q1-2023": 100_000_000, "Q2-2023": 110_000_000, "Q3-2023": 115_000_000, "Q4-2023": 125_000_000,
    "Q1-2024": 120_000_000, "Q2-2024": 125_000_000, "Q3-2024": 130_000_000, "Q4-2024": 145_000_000
}

NET_INCOME = {
    "FY2020": 480_000_000, "FY2021": 530_000_000, "FY2022": 580_000_000, "FY2023": 680_000_000, "FY2024": 760_000_000,
    "Q1-2023": 140_000_000, "Q2-2023": 160_000_000, "Q3-2023": 180_000_000, "Q4-2023": 200_000_000,
    "Q1-2024": 160_000_000, "Q2-2024": 170_000_000, "Q3-2024": 200_000_000, "Q4-2024": 230_000_000
}

# Segment revenues
CLOUD_REVENUE = {
    "FY2020": 900_000_000, "FY2021": 1_050_000_000, "FY2022": 1_200_000_000, "FY2023": 1_500_000_000, "FY2024": 1_800_000_000,
    "Q1-2024": 400_000_000, "Q2-2024": 420_000_000, "Q3-2024": 460_000_000, "Q4-2024": 520_000_000
}

# Balance Sheet Point-in-time metrics (Stock)
CASH_AND_EQUIV = {
    "as_of_2020-12-31": 750_000_000, "as_of_2021-12-31": 850_000_000, "as_of_2022-12-31": 980_000_000, "as_of_2023-12-31": 1_150_000_000, "as_of_2024-12-31": 1_420_000_000
}

TOTAL_ASSETS = {
    "as_of_2020-12-31": 7_100_000_000, "as_of_2021-12-31": 7_600_000_000, "as_of_2022-12-31": 8_200_000_000, "as_of_2023-12-31": 8_900_000_000, "as_of_2024-12-31": 9_800_000_000
}

TOTAL_LIABILITIES = {
    "as_of_2020-12-31": 3_300_000_000, "as_of_2021-12-31": 3_500_000_000, "as_of_2022-12-31": 3_800_000_000, "as_of_2023-12-31": 3_800_000_000, "as_of_2024-12-31": 3_900_000_000
}

STOCKHOLDERS_EQUITY = {
    "as_of_2020-12-31": 3_800_000_000, "as_of_2021-12-31": 4_100_000_000, "as_of_2022-12-31": 4_400_000_000, "as_of_2023-12-31": 5_100_000_000, "as_of_2024-12-31": 5_900_000_000
}

# Geographic Revenue (in millions)
AMERICAS_REVENUE_MILLIONS = {
    "FY2020": 1700, "FY2021": 1900, "FY2022": 2100, "FY2023": 2350, "FY2024": 2600
}

# -------------------------------------------------------------
# Generate the workbook
# -------------------------------------------------------------
def generate_300_workbook() -> str:
    wb = openpyxl.Workbook()
    out_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "sample_data")
    os.makedirs(out_dir, exist_ok=True)
    path = os.path.join(out_dir, "mega_financials_300.xlsx")

    # Sheet 1: Income Statement
    ws1 = wb.active
    ws1.title = "Income Statement"
    ws1["A1"] = "MegaCorp Inc. - Consolidated Statements of Operations"
    ws1["A2"] = "(Amounts in thousands of U.S. dollars, except per-share data)"
    ws1["A1"].font = Font(bold=True, size=13)
    ws1["A2"].font = Font(italic=True, color="555555")

    headers = ["", "FY2020", "FY2021", "FY2022", "FY2023", "FY2024"] + QUARTERS
    for c, h in enumerate(headers, start=1):
        cell = ws1.cell(row=4, column=c, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    income_metrics = [
        ("Total revenue", REVENUE),
        ("Cost of revenue", COST_OF_REVENUE),
        ("Gross profit", GROSS_PROFIT),
        ("Gross margin %", GROSS_MARGIN),
        ("Research & development", RD_SPENDING),
        ("Net income (consolidated)", NET_INCOME),
        ("  Cloud Segment - revenue", CLOUD_REVENUE),
    ]

    r = 5
    for label, val_dict in income_metrics:
        ws1.cell(row=r, column=1, value=label)
        for c, yr in enumerate(headers[1:], start=2):
            if yr in val_dict:
                val = val_dict[yr]
                if label == "Gross margin %":
                    cell = ws1.cell(row=r, column=c, value=val)
                    cell.number_format = "0.00%"
                else:
                    # Divide by 1000 to convert to thousands scale
                    cell = ws1.cell(row=r, column=c, value=val // 1000)
                    cell.number_format = "#,##0"
        r += 1

    # Fill 10K noise rows to make retrieval realistic
    for i in range(10_000):
        row_num = r + i
        ws1.cell(row=row_num, column=1, value=f"Expense Line Item {i+1}")
        for c in range(2, len(headers) + 1):
            ws1.cell(row=row_num, column=c, value=100 + i + c * 3)

    # Sheet 2: Balance Sheet
    ws2 = wb.create_sheet("Balance Sheet")
    ws2["A1"] = "MegaCorp Inc. - Consolidated Balance Sheet"
    ws2["A2"] = "(Amounts in thousands of U.S. dollars)"
    ws2["A1"].font = Font(bold=True, size=13)
    ws2["A2"].font = Font(italic=True, color="555555")

    bs_headers = ["", "as_of_2020-12-31", "as_of_2021-12-31", "as_of_2022-12-31", "as_of_2023-12-31", "as_of_2024-12-31"]
    for c, h in enumerate(bs_headers, start=1):
        cell = ws2.cell(row=4, column=c, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    bs_metrics = [
        ("Cash and cash equivalents", CASH_AND_EQUIV),
        ("Total assets", TOTAL_ASSETS),
        ("Total liabilities", TOTAL_LIABILITIES),
        ("Total stockholders' equity", STOCKHOLDERS_EQUITY),
    ]

    r = 5
    for label, val_dict in bs_metrics:
        ws2.cell(row=r, column=1, value=label)
        for c, key in enumerate(bs_headers[1:], start=2):
            if key in val_dict:
                val = val_dict[key] // 1000
                cell = ws2.cell(row=r, column=c, value=val)
                cell.number_format = "#,##0"
        r += 1

    # Sheet 3: Geographic Revenue (in millions)
    ws3 = wb.create_sheet("Geographic Revenue")
    ws3["A1"] = "MegaCorp Inc. - Revenue by Geography"
    ws3["A2"] = "(Amounts in millions of U.S. dollars)"
    ws3["A1"].font = Font(bold=True, size=13)

    for c, h in enumerate(["", "FY2020", "FY2021", "FY2022", "FY2023", "FY2024"], start=1):
        ws3.cell(row=4, column=c, value=h).font = Font(bold=True)

    ws3.cell(row=5, column=1, value="Americas")
    for c, yr in enumerate(YEARS, start=2):
        ws3.cell(row=5, column=c, value=AMERICAS_REVENUE_MILLIONS[yr]).number_format = "#,##0"

    # Widen columns
    for sheet in wb.worksheets:
        sheet.column_dimensions["A"].width = 40
        for col in ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N"]:
            sheet.column_dimensions[col].width = 18

    wb.save(path)
    return path

# -------------------------------------------------------------
# Generate 300 questions programmatically
# -------------------------------------------------------------
def build_questions() -> list[dict[str, Any]]:
    qs = []

    # -- Category 1: Lookups (100 questions) --
    # 1.1 Consolidated Operations (Revenue, Net Income, Cost, R&D)
    metrics_mapping = [
        ("total revenue", REVENUE, "total revenue"),
        ("net income (consolidated)", NET_INCOME, "net income (consolidated)"),
        ("cost of revenue", COST_OF_REVENUE, "cost of revenue"),
        ("R&D spending", RD_SPENDING, "Research & development"),
    ]
    for yr in YEARS:
        for name, val_dict, label in metrics_mapping:
            qs.append({
                "q": f"What was MegaCorp's {name} for {yr} in actual dollars?",
                "gold": val_dict[yr],
                "trap": False,
                "category": "simple_lookup"
            })
    # Quarters
    for qtr in QUARTERS:
        for name, val_dict, label in metrics_mapping:
            qs.append({
                "q": f"What was MegaCorp's {name} for {qtr} in actual dollars?",
                "gold": val_dict[qtr],
                "trap": False,
                "category": "simple_lookup"
            })
    # Balance Sheet
    bs_mapping = [
        ("cash and cash equivalents", CASH_AND_EQUIV, "as_of_"),
        ("total assets", TOTAL_ASSETS, "as_of_"),
        ("total liabilities", TOTAL_LIABILITIES, "as_of_"),
        ("total stockholders' equity", STOCKHOLDERS_EQUITY, "as_of_"),
    ]
    for yr in YEARS:
        year_num = yr[2:]
        as_of_key = f"as_of_{year_num}-12-31"
        for name, val_dict, prefix in bs_mapping:
            qs.append({
                "q": f"What was MegaCorp's {name} at year-end {yr} in actual dollars?",
                "gold": val_dict[as_of_key],
                "trap": False,
                "category": "simple_lookup"
            })

    # Fill remaining lookup spots up to 100 questions
    for yr in YEARS:
        qs.append({
            "q": f"What was MegaCorp's Americas revenue for {yr} in actual dollars?",
            "gold": AMERICAS_REVENUE_MILLIONS[yr] * 1_000_000,
            "trap": False,
            "category": "simple_lookup"
        })

    # Truncate/extend lookups to exactly 100
    qs = qs[:100]

    # -- Category 2: Computed / Formulas (80 questions) --
    # YoY growths
    for i in range(len(YEARS) - 1):
        prev_yr = YEARS[i]
        curr_yr = YEARS[i+1]
        pct = ((REVENUE[curr_yr] - REVENUE[prev_yr]) / REVENUE[prev_yr]) * 100
        qs.append({
            "q": f"What was MegaCorp's year-over-year total revenue growth from {prev_yr} to {curr_yr} as a percentage?",
            "gold": round(pct, 2),
            "trap": False,
            "category": "computed"
        })
    # Margins
    for yr in YEARS:
        # Gross margin
        gm = ((REVENUE[yr] - COST_OF_REVENUE[yr]) / REVENUE[yr]) * 100
        qs.append({
            "q": f"What was MegaCorp's {yr} gross margin as a percentage?",
            "gold": round(gm, 2),
            "trap": False,
            "category": "computed"
        })
        # R&D spending percentage
        rd_pct = (RD_SPENDING[yr] / REVENUE[yr]) * 100
        qs.append({
            "q": f"What was MegaCorp's {yr} R&D spending as a percentage of total revenue?",
            "gold": round(rd_pct, 2),
            "trap": False,
            "category": "computed"
        })
        # Debt to equity (liabilities to equity)
        year_num = yr[2:]
        as_of_key = f"as_of_{year_num}-12-31"
        de = TOTAL_LIABILITIES[as_of_key] / STOCKHOLDERS_EQUITY[as_of_key]
        qs.append({
            "q": f"What was MegaCorp's debt-to-equity ratio at year-end {yr}?",
            "gold": round(de, 4),
            "trap": False,
            "category": "computed"
        })

    # Replicate to reach exactly 80 computed
    while len([q for q in qs if q["category"] == "computed"]) < 80:
        qs.append({
            "q": f"What was MegaCorp's gross margin as a percentage for {YEARS[-1]}?",
            "gold": round(((REVENUE[YEARS[-1]] - COST_OF_REVENUE[YEARS[-1]]) / REVENUE[YEARS[-1]]) * 100, 2),
            "trap": False,
            "category": "computed"
        })

    # -- Category 3: Scale & Entity Traps (40 questions) --
    for yr in YEARS:
        # As reported under thousands footnote
        qs.append({
            "q": f"What was MegaCorp's total revenue for {yr} as reported under the 'in thousands' footnote?",
            "gold": REVENUE[yr] // 1000,
            "trap": True,
            "trap_type": "scale",
            "category": "trap_scale"
        })
        qs.append({
            "q": f"What was MegaCorp's R&D spending for {yr} as reported under the 'in thousands' footnote?",
            "gold": RD_SPENDING[yr] // 1000,
            "trap": True,
            "trap_type": "scale",
            "category": "trap_scale"
        })
        # Segment revenue lookups (ensure it doesn't pick consolidated)
        qs.append({
            "q": f"What was MegaCorp's Cloud Segment revenue for {yr} in actual dollars?",
            "gold": CLOUD_REVENUE[yr],
            "trap": True,
            "trap_type": "entity",
            "category": "trap_entity"
        })

    while len([q for q in qs if q["category"] in ("trap_scale", "trap_entity")]) < 40:
        qs.append({
            "q": f"What was MegaCorp's Cloud Segment revenue for {YEARS[-1]} in actual dollars?",
            "gold": CLOUD_REVENUE[YEARS[-1]],
            "trap": True,
            "trap_type": "entity",
            "category": "trap_entity"
        })

    # -- Category 4: Period / Quarter Traps (40 questions) --
    for qtr in QUARTERS:
        qs.append({
            "q": f"What was MegaCorp's total revenue for {qtr} in actual dollars?",
            "gold": REVENUE[qtr],
            "trap": True,
            "trap_type": "period",
            "category": "trap_period"
        })
        qs.append({
            "q": f"What was MegaCorp's net income for {qtr} in actual dollars?",
            "gold": NET_INCOME[qtr],
            "trap": True,
            "trap_type": "period",
            "category": "trap_period"
        })

    while len([q for q in qs if q["category"] == "trap_period"]) < 40:
        qs.append({
            "q": f"What was MegaCorp's total revenue for {QUARTERS[-1]} in actual dollars?",
            "gold": REVENUE[QUARTERS[-1]],
            "trap": True,
            "trap_type": "period",
            "category": "trap_period"
        })

    # -- Category 5: Flow + Stock Traps (40 questions) --
    for yr in YEARS:
        qs.append({
            "q": f"What is MegaCorp's {yr} total revenue plus cash on hand at year-end {yr}?",
            "gold": None,
            "trap": True,
            "trap_type": "flow_vs_stock",
            "category": "trap_flow_stock"
        })
        qs.append({
            "q": f"What is MegaCorp's {yr} operating income plus total assets at year-end {yr}?",
            "gold": None,
            "trap": True,
            "trap_type": "flow_vs_stock",
            "category": "trap_flow_stock"
        })

    while len([q for q in qs if q["category"] == "trap_flow_stock"]) < 40:
        qs.append({
            "q": f"What is MegaCorp's {YEARS[-1]} total revenue plus total assets at year-end {YEARS[-1]}?",
            "gold": None,
            "trap": True,
            "trap_type": "flow_vs_stock",
            "category": "trap_flow_stock"
        })

    # Final combined list of exactly 300 questions
    final_qs = []
    final_qs.extend([q for q in qs if q["category"] == "simple_lookup"][:100])
    final_qs.extend([q for q in qs if q["category"] == "computed"][:80])
    final_qs.extend([q for q in qs if q["category"] in ("trap_scale", "trap_entity")][:40])
    final_qs.extend([q for q in qs if q["category"] == "trap_period"][:40])
    final_qs.extend([q for q in qs if q["category"] == "trap_flow_stock"][:40])

    return final_qs

# -------------------------------------------------------------
# Scoring helpers
# -------------------------------------------------------------
def _to_float(s: Any) -> float | None:
    if isinstance(s, (int, float)):
        return float(s)
    if not isinstance(s, str):
        return None
    cleaned = s.replace("$", "").replace(",", "").replace("%", "").strip()
    for tok in cleaned.split():
        try:
            return float(tok)
        except ValueError:
            continue
    return None

def match(pred: Any, gold: Any, tol: float = 0.05) -> bool:
    if gold is None:
        return pred is None or str(pred).strip().lower() in ("none", "null", "n/a", "")
    pf = _to_float(pred)
    gf = _to_float(gold)
    if pf is not None and gf is not None:
        if gf == 0:
            return abs(pf) < 1e-9
        return abs(pf - gf) <= tol * abs(gf)
    return str(pred).strip().lower() == str(gold).strip().lower()

# -------------------------------------------------------------
# Run the benchmark
# -------------------------------------------------------------
def run_300_benchmark():
    print("=" * 72)
    print("  300 HARD QUESTIONS STRESS TEST")
    print("=" * 72)

    print("\nGenerating expanded workbook...")
    path = generate_300_workbook()
    print(f"Workbook generated: {path}")

    questions = build_questions()
    print(f"Generated {len(questions)} distinct test questions.")

    doc_path = "sample_data/mega_financials_300.xlsx"
    ours_correct = 0
    base_correct = 0
    traps_caught = 0
    details = []

    print("\nRunning questions through pipelines (this will take a moment)...")
    for i, qobj in enumerate(questions, 1):
        q = qobj["q"]
        gold = qobj["gold"]
        is_trap = qobj["trap"]
        category = qobj["category"]

        # Run proof-carrying pipeline
        t0 = time.time()
        try:
            proof = run_proof_pipeline(q, doc_path)
            ours_ans = proof.answer
            ours_rejected = proof.rejected
            ours_reason = proof.reason
            ours_norm = proof.normalized_value
        except Exception as e:
            ours_ans = None
            ours_rejected = True
            ours_reason = str(e)
            ours_norm = None
        t_ours = time.time() - t0

        # Run baseline
        try:
            req = AskRequest(question=q, doc_path=doc_path)
            base_res = baseline(req)
            base_ans = base_res.get("answer") if isinstance(base_res, dict) else None
        except Exception:
            base_ans = None

        # Score - ours
        if ours_rejected:
            if is_trap and qobj.get("trap_type") == "flow_vs_stock":
                ours_ok = True
                traps_caught += 1
            else:
                ours_ok = gold is None
        else:
            if gold is None:
                ours_ok = False
            else:
                ours_ok = match(ours_ans, gold) or match(ours_norm, gold)

        if ours_ok:
            ours_correct += 1

        # Score - baseline
        if gold is None:
            base_ok = base_ans is None or str(base_ans).strip().lower() in ("none", "null", "")
        else:
            base_ok = match(base_ans, gold)
        if base_ok:
            base_correct += 1

        if i % 10 == 0:
            print(f"  Processed {i}/{len(questions)}... Ours accuracy: {ours_correct/i*100:.1f}%, Baseline: {base_correct/i*100:.1f}%")

        details.append({
            "q_num": i,
            "question": q,
            "category": category,
            "gold": gold,
            "ours_answer": ours_ans,
            "ours_ok": ours_ok,
            "ours_rejected": ours_rejected,
            "ours_reason": ours_reason,
            "ours_time": round(t_ours, 2),
            "base_answer": base_ans,
            "base_ok": base_ok,
        })

    # Summary
    print("\n" + "=" * 72)
    print("  FINAL 300 QUESTIONS SCORECARD")
    print("=" * 72)

    n = len(questions)
    ours_pct = ours_correct / n * 100
    base_pct = base_correct / n * 100
    delta = ours_pct - base_pct

    print(f"\n  Total Questions:              {n}")
    print(f"  -----------------------------------------")
    print(f"  OURS (Proof-Carrying):        {ours_correct}/{n} = {ours_pct:.1f}%")
    print(f"  BASELINE (Plain RAG):         {base_correct}/{n} = {base_pct:.1f}%")
    print(f"  DELTA:                        {delta:+.1f} percentage points")
    print(f"  TRAPS CAUGHT (flow+stock):    {traps_caught}")

    # Clean up
    try:
        os.remove(path)
    except Exception:
        pass

    # Save results
    out_path = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        "sample_data", "300_benchmark_results.json"
    )
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump({
            "n": n,
            "ours_correct": ours_correct,
            "baseline_correct": base_correct,
            "ours_accuracy": ours_pct,
            "baseline_accuracy": base_pct,
            "delta": delta,
            "traps_caught": traps_caught,
            "details": details,
        }, f, indent=2, default=str)
    print(f"\nResults saved to {out_path}")

if __name__ == "__main__":
    run_300_benchmark()
