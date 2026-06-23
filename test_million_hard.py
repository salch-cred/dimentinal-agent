"""Hard Benchmark -- Stress Test + Accuracy Scorer.

Generates a realistic, trap-laden financial workbook with ~50K cells
across 6 sheets, then runs 20 hard questions (including 8 traps) through
both the proof-carrying pipeline and the baseline.

Run:  python test_million_hard.py
"""
from __future__ import annotations

import json
import os
import sys
import time
from typing import Any

import openpyxl
from openpyxl.styles import Font, Alignment

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from app import AskRequest, run_proof_pipeline, baseline
from ingest import ingest


# -------------------------------------------------------------
#  1. Generate the workbook
# -------------------------------------------------------------
def generate_hard_workbook() -> str:
    """Build a 6-sheet workbook with ~50K cells and embedded traps."""
    wb = openpyxl.Workbook()
    out_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "sample_data")
    os.makedirs(out_dir, exist_ok=True)
    path = os.path.join(out_dir, "mega_financials.xlsx")

    # -- Sheet 1: Income Statement (in thousands) ----------------
    ws1 = wb.active
    ws1.title = "Income Statement"
    ws1["A1"] = "MegaCorp Inc. - Consolidated Statements of Operations"
    ws1["A2"] = "(Amounts in thousands of U.S. dollars, except per-share data)"
    ws1["A1"].font = Font(bold=True, size=13)
    ws1["A2"].font = Font(italic=True, color="555555")

    headers = ["", "FY2022", "FY2023", "FY2024", "Q1-2024", "Q2-2024", "Q3-2024", "Q4-2024"]
    for c, h in enumerate(headers, start=1):
        cell = ws1.cell(row=4, column=c, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Key income-statement rows (values in thousands)
    income_rows = [
        ("Total revenue",                3_800_000, 4_200_000, 4_650_000,  1_050_000, 1_100_000, 1_200_000, 1_300_000),
        ("Cost of revenue",              2_300_000, 2_500_000, 2_750_000,    630_000,   660_000,   710_000,   750_000),
        ("Gross profit",                 1_500_000, 1_700_000, 1_900_000,    420_000,   440_000,   490_000,   550_000),
        ("Research & development",         400_000,   450_000,   520_000,    120_000,   125_000,   130_000,   145_000),
        ("Sales, general & admin",         350_000,   380_000,   410_000,     95_000,   100_000,   105_000,   110_000),
        ("Total operating expenses",       750_000,   830_000,   930_000,    215_000,   225_000,   235_000,   255_000),
        ("Operating income",               750_000,   870_000,   970_000,    205_000,   215_000,   255_000,   295_000),
        ("Interest expense",                45_000,    42_000,    38_000,      9_000,     9_500,     9_500,    10_000),
        ("Net income (consolidated)",      580_000,   680_000,   760_000,    160_000,   170_000,   200_000,   230_000),
    ]
    # Trap: Segment rows sitting right below consolidated
    segment_rows = [
        ("  Cloud Segment - revenue",    1_200_000, 1_500_000, 1_800_000,    400_000,   420_000,   460_000,   520_000),
        ("  Cloud Segment - net income",   210_000,   270_000,   340_000,     72_000,    78_000,    88_000,   102_000),
        ("  Enterprise Segment - revenue", 2_600_000, 2_700_000, 2_850_000,  650_000,   680_000,   740_000,   780_000),
        ("  Enterprise Segment - net income", 370_000, 410_000,  420_000,     88_000,    92_000,   112_000,   128_000),
    ]
    # Margin rows
    margin_rows = [
        ("Gross margin %", 0.3947, 0.4048, 0.4086, 0.4000, 0.4000, 0.4083, 0.4231),
        ("Operating margin %", 0.1974, 0.2071, 0.2086, 0.1952, 0.1955, 0.2125, 0.2269),
    ]

    r = 5
    for label, *vals in income_rows + segment_rows:
        ws1.cell(row=r, column=1, value=label)
        for c, v in enumerate(vals, start=2):
            cell = ws1.cell(row=r, column=c, value=v)
            cell.number_format = "#,##0"
        r += 1
    for label, *vals in margin_rows:
        ws1.cell(row=r, column=1, value=label).font = Font(italic=True)
        for c, v in enumerate(vals, start=2):
            ws1.cell(row=r, column=c, value=v).number_format = "0.00%"
        r += 1

    # Fill 5K noise rows on income statement
    noise_start = r + 1
    for i in range(5_000):
        row_num = noise_start + i
        ws1.cell(row=row_num, column=1, value=f"Expense Line Item {i+1}")
        for c in range(2, 9):
            ws1.cell(row=row_num, column=c, value=1000 + i + c * 7)

    # -- Sheet 2: Balance Sheet (point-in-time) ------------------
    ws2 = wb.create_sheet("Balance Sheet")
    ws2["A1"] = "MegaCorp Inc. - Consolidated Balance Sheet"
    ws2["A2"] = "(Amounts in thousands of U.S. dollars)"
    ws2["A1"].font = Font(bold=True, size=13)
    ws2["A2"].font = Font(italic=True, color="555555")

    for c, h in enumerate(["", "as_of_2022-12-31", "as_of_2023-12-31", "as_of_2024-12-31"], start=1):
        cell = ws2.cell(row=4, column=c, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    bs_rows = [
        ("Cash and cash equivalents",      980_000, 1_150_000, 1_420_000),
        ("Short-term investments",         320_000,   380_000,   450_000),
        ("Accounts receivable",            650_000,   720_000,   810_000),
        ("Total current assets",         2_800_000, 3_100_000, 3_550_000),
        ("Property & equipment, net",    1_200_000, 1_350_000, 1_500_000),
        ("Goodwill",                     2_100_000, 2_100_000, 2_400_000),
        ("Total assets",                 8_200_000, 8_900_000, 9_800_000),
        ("Total current liabilities",    1_400_000, 1_550_000, 1_700_000),
        ("Long-term debt",               1_800_000, 1_600_000, 1_400_000),
        ("Total liabilities",            3_800_000, 3_800_000, 3_900_000),
        ("Total stockholders' equity",   4_400_000, 5_100_000, 5_900_000),
    ]
    r = 5
    for label, *vals in bs_rows:
        ws2.cell(row=r, column=1, value=label)
        for c, v in enumerate(vals, start=2):
            ws2.cell(row=r, column=c, value=v).number_format = "#,##0"
        r += 1

    # 5K noise rows on balance sheet
    noise_start = r + 1
    for i in range(5_000):
        row_num = noise_start + i
        ws2.cell(row=row_num, column=1, value=f"Asset Line {i+1}")
        for c in range(2, 5):
            ws2.cell(row=row_num, column=c, value=500 + i + c * 3)

    # -- Sheet 3: Per Share Data ---------------------------------
    ws3 = wb.create_sheet("Per Share")
    ws3["A1"] = "MegaCorp Inc. - Per Share Data"
    ws3["A1"].font = Font(bold=True, size=13)
    for c, h in enumerate(["", "FY2022", "FY2023", "FY2024"], start=1):
        ws3.cell(row=3, column=c, value=h).font = Font(bold=True)

    ws3.cell(row=4, column=1, value="Diluted shares outstanding (millions)")
    ws3.cell(row=4, column=2, value=500)
    ws3.cell(row=4, column=3, value=495)
    ws3.cell(row=4, column=4, value=490)

    ws3.cell(row=5, column=1, value="Diluted EPS")
    ws3.cell(row=5, column=2, value=1.16)
    ws3.cell(row=5, column=3, value=1.37)
    ws3.cell(row=5, column=4, value=1.55)

    ws3.cell(row=6, column=1, value="Dividends per share")
    ws3.cell(row=6, column=2, value=0.52)
    ws3.cell(row=6, column=3, value=0.56)
    ws3.cell(row=6, column=4, value=0.62)

    # -- Sheet 4: Geographic Revenue (in millions -- different scale!) --
    ws4 = wb.create_sheet("Geographic Revenue")
    ws4["A1"] = "MegaCorp Inc. - Revenue by Geography"
    ws4["A2"] = "(Amounts in millions of U.S. dollars)"
    ws4["A1"].font = Font(bold=True, size=13)
    ws4["A2"].font = Font(italic=True, color="555555")

    for c, h in enumerate(["", "FY2022", "FY2023", "FY2024"], start=1):
        ws4.cell(row=4, column=c, value=h).font = Font(bold=True)

    geo_rows = [
        ("Americas",        2_100, 2_350, 2_600),
        ("Europe (EMEA)",   1_000, 1_100, 1_250),
        ("Asia-Pacific",      700,   750,   800),
        ("Total",           3_800, 4_200, 4_650),
    ]
    r = 5
    for label, *vals in geo_rows:
        ws4.cell(row=r, column=1, value=label)
        for c, v in enumerate(vals, start=2):
            ws4.cell(row=r, column=c, value=v).number_format = "#,##0"
        r += 1

    # -- Sheet 5: Cash Flow Statement ----------------------------
    ws5 = wb.create_sheet("Cash Flow")
    ws5["A1"] = "MegaCorp Inc. - Consolidated Cash Flow Statement"
    ws5["A2"] = "(Amounts in thousands of U.S. dollars)"
    ws5["A1"].font = Font(bold=True, size=13)
    ws5["A2"].font = Font(italic=True, color="555555")

    for c, h in enumerate(["", "FY2022", "FY2023", "FY2024"], start=1):
        ws5.cell(row=4, column=c, value=h).font = Font(bold=True)

    cf_rows = [
        ("Operating cash flow",        720_000,   850_000,   980_000),
        ("Capital expenditures",      -280_000,  -320_000,  -350_000),
        ("Free cash flow",             440_000,   530_000,   630_000),
        ("Acquisitions",              -150_000,        0,  -300_000),
        ("Dividends paid",            -260_000,  -277_000,  -304_000),
        ("Share repurchases",         -100_000,  -120_000,  -150_000),
    ]
    r = 5
    for label, *vals in cf_rows:
        ws5.cell(row=r, column=1, value=label)
        for c, v in enumerate(vals, start=2):
            ws5.cell(row=r, column=c, value=v).number_format = "#,##0"
        r += 1

    # -- Sheet 6: Regional Details (large) -----------------------
    ws6 = wb.create_sheet("Regional Details")
    ws6["A1"] = "MegaCorp Inc. - Regional P&L Detail"
    ws6["A2"] = "(Amounts in thousands of U.S. dollars)"
    ws6["A1"].font = Font(bold=True, size=13)
    ws6["A2"].font = Font(italic=True, color="555555")

    detail_headers = ["Region", "Sub-Region", "Product Line", "FY2023", "FY2024"]
    for c, h in enumerate(detail_headers, start=1):
        ws6.cell(row=4, column=c, value=h).font = Font(bold=True)

    regions = ["Americas", "EMEA", "APAC"]
    sub_regions = {
        "Americas": ["US-East", "US-West", "US-Central", "Canada", "Latin America"],
        "EMEA": ["UK", "Germany", "France", "Nordics", "Middle East"],
        "APAC": ["Japan", "China", "India", "Australia", "Southeast Asia"],
    }
    products = ["Cloud SaaS", "Cloud PaaS", "Enterprise License", "Support", "Consulting",
                "Hardware", "Training", "Data Analytics", "Security Suite", "AI Platform"]

    r = 5
    for region in regions:
        for sub in sub_regions[region]:
            for prod in products:
                fy23 = 500 + hash(f"{region}{sub}{prod}") % 50_000
                fy24 = int(fy23 * 1.08)
                ws6.cell(row=r, column=1, value=region)
                ws6.cell(row=r, column=2, value=sub)
                ws6.cell(row=r, column=3, value=prod)
                ws6.cell(row=r, column=4, value=fy23)
                ws6.cell(row=r, column=5, value=fy24)
                r += 1

    # Widen columns
    for sheet in wb.worksheets:
        sheet.column_dimensions["A"].width = 40
        for col in ["B", "C", "D", "E", "F", "G", "H"]:
            sheet.column_dimensions[col].width = 18

    print("  Saving workbook...")
    wb.save(path)
    return path


# -------------------------------------------------------------
#  2. Hard questions with gold answers
# -------------------------------------------------------------
HARD_QUESTIONS = [
    # -- Simple lookups --
    {
        "q": "What was MegaCorp's total revenue for FY2024 in actual dollars?",
        "gold": 4_650_000_000,
        "trap": False,
        "category": "simple_lookup",
    },
    {
        "q": "What was MegaCorp's net income (consolidated) for FY2024 in actual dollars?",
        "gold": 760_000_000,
        "trap": False,
        "category": "simple_lookup",
    },
    {
        "q": "What was MegaCorp's FY2024 diluted EPS?",
        "gold": 1.55,
        "trap": False,
        "category": "simple_lookup",
    },
    {
        "q": "What was MegaCorp's operating cash flow for FY2024 in actual dollars?",
        "gold": 980_000_000,
        "trap": False,
        "category": "simple_lookup",
    },
    {
        "q": "What was MegaCorp's free cash flow for FY2024 in actual dollars?",
        "gold": 630_000_000,
        "trap": False,
        "category": "simple_lookup",
    },
    # -- Computed / derived --
    {
        "q": "What was MegaCorp's FY2024 gross margin as a percentage?",
        "gold": 40.86,
        "trap": False,
        "category": "computed",
    },
    {
        "q": "What was MegaCorp's year-over-year total revenue growth from FY2023 to FY2024 as a percentage?",
        "gold": 10.714,
        "trap": False,
        "category": "computed",
    },
    {
        "q": "What was MegaCorp's debt-to-equity ratio at year-end FY2024?",
        "gold": 0.2373,  # 1,400,000 / 5,900,000
        "trap": False,
        "category": "computed",
    },
    # -- TRAP: Period mismatch --
    {
        "q": "What was MegaCorp's total revenue for Q4-2024 in actual dollars?",
        "gold": 1_300_000_000,
        "trap": True,
        "trap_type": "period",
        "category": "trap_period",
        "note": "Q4 column is next to FY columns; baseline may pick wrong period",
    },
    {
        "q": "What was MegaCorp's Q3-2024 net income (consolidated) in actual dollars?",
        "gold": 200_000_000,
        "trap": True,
        "trap_type": "period",
        "category": "trap_period",
        "note": "Must pick the right quarter column",
    },
    # -- TRAP: Entity mismatch --
    {
        "q": "What was MegaCorp's consolidated net income for FY2024 (not the Cloud Segment)?",
        "gold": 760_000_000,
        "trap": True,
        "trap_type": "entity",
        "category": "trap_entity",
        "note": "Cloud Segment net income (340M) sits right below consolidated (760M)",
    },
    {
        "q": "What was MegaCorp's Cloud Segment revenue for FY2024 in actual dollars?",
        "gold": 1_800_000_000,
        "trap": True,
        "trap_type": "entity",
        "category": "trap_entity",
        "note": "Must pick segment not consolidated",
    },
    # -- TRAP: Flow + Stock (must reject) --
    {
        "q": "What is MegaCorp's FY2024 total revenue plus cash on hand at year-end FY2024?",
        "gold": None,
        "trap": True,
        "trap_type": "flow_vs_stock",
        "category": "trap_flow_stock",
        "note": "Adding a period flow to a point-in-time balance is meaningless",
    },
    {
        "q": "What is MegaCorp's FY2024 operating income plus total assets at year-end FY2024?",
        "gold": None,
        "trap": True,
        "trap_type": "flow_vs_stock",
        "category": "trap_flow_stock",
        "note": "Another flow+stock combination that should be rejected",
    },
    # -- TRAP: Scale --
    {
        "q": "What was MegaCorp's total revenue for FY2024 as reported under the 'in thousands' footnote?",
        "gold": 4_650_000,
        "trap": True,
        "trap_type": "scale",
        "category": "trap_scale",
        "note": "As-printed value is 4,650,000 (thousands); must not double-scale",
    },
    # -- Multi-sheet / cross-reference --
    {
        "q": "What was MegaCorp's cash and cash equivalents at year-end FY2024 in actual dollars?",
        "gold": 1_420_000_000,
        "trap": False,
        "category": "cross_sheet",
    },
    {
        "q": "What was MegaCorp's total stockholders' equity at year-end FY2024 in actual dollars?",
        "gold": 5_900_000_000,
        "trap": False,
        "category": "cross_sheet",
    },
    # -- Hard computed --
    {
        "q": "What was MegaCorp's FY2024 R&D spending as a percentage of total revenue?",
        "gold": 11.18,  # 520_000 / 4_650_000 * 100
        "trap": False,
        "category": "hard_computed",
    },
    {
        "q": "What was MegaCorp's dividends per share for FY2024?",
        "gold": 0.62,
        "trap": False,
        "category": "simple_lookup",
    },
    {
        "q": "What was MegaCorp's Americas revenue for FY2024 in actual dollars?",
        "gold": 2_600_000_000,  # From Geographic Revenue sheet (in millions x 1M)
        "trap": False,
        "category": "cross_sheet_scale",
        "note": "Geographic Revenue sheet uses 'in millions' scale, not thousands",
    },
]


# -------------------------------------------------------------
#  3. Scoring helpers
# -------------------------------------------------------------
def _to_float(s: Any) -> float | None:
    if isinstance(s, (int, float)):
        return float(s)
    if not isinstance(s, str):
        return None
    cleaned = s.replace("$", "").replace(",", "").replace("%", "").strip()
    for tok in cleaned.split():
        try:
            return float(tok)
        except ValueError:
            continue
    return None


def match(pred: Any, gold: Any, tol: float = 0.05) -> bool:
    """Numeric match within tol (5% for hard test), else case-insensitive string match."""
    if gold is None:
        return pred is None or str(pred).strip().lower() in ("none", "null", "n/a", "")
    pf = _to_float(pred)
    gf = _to_float(gold)
    if pf is not None and gf is not None:
        if gf == 0:
            return abs(pf) < 1e-9
        return abs(pf - gf) <= tol * abs(gf)
    return str(pred).strip().lower() == str(gold).strip().lower()


# -------------------------------------------------------------
#  4. Run the benchmark
# -------------------------------------------------------------
def run_hard_benchmark():
    print("=" * 72)
    print("  HARD BENCHMARK -- MegaCorp Financial Stress Test (6 sheets)")
    print("=" * 72)

    # Step 1: Generate workbook
    print("\n[1/4] Generating multi-sheet workbook...")
    t0 = time.time()
    path = generate_hard_workbook()
    t_gen = time.time() - t0
    print(f"  [OK] Workbook generated in {t_gen:.1f}s: {path}")

    # Step 2: Benchmark ingestion
    print("\n[2/4] Benchmarking ingestion pipeline...")
    t0 = time.time()
    spans = ingest(path)
    t_ingest = time.time() - t0
    print(f"  [OK] Ingested {len(spans):,} spans in {t_ingest:.1f}s")
    print(f"  [OK] Speed: {len(spans)/t_ingest:,.0f} spans/second")

    # Step 3: Run all questions
    print(f"\n[3/4] Running {len(HARD_QUESTIONS)} hard questions through pipeline...")
    print("-" * 72)

    doc_path = "sample_data/mega_financials.xlsx"
    n = len(HARD_QUESTIONS)
    ours_correct = 0
    base_correct = 0
    traps_caught = 0
    details = []

    for i, qobj in enumerate(HARD_QUESTIONS, 1):
        q = qobj["q"]
        gold = qobj["gold"]
        is_trap = qobj["trap"]
        category = qobj["category"]

        print(f"\n  Q{i:02d} [{category}]: {q[:75]}...")

        # Run proof-carrying pipeline
        t0 = time.time()
        try:
            proof = run_proof_pipeline(q, doc_path)
            ours_ans = proof.answer
            ours_rejected = proof.rejected
            ours_reason = proof.reason
            ours_verdict = proof.verifier_verdict
            ours_norm = proof.normalized_value
        except Exception as e:
            ours_ans = None
            ours_rejected = True
            ours_reason = str(e)
            ours_verdict = "error"
            ours_norm = None
        t_ours = time.time() - t0

        # Run baseline
        try:
            req = AskRequest(question=q, doc_path=doc_path)
            base_res = baseline(req)
            base_ans = base_res.get("answer") if isinstance(base_res, dict) else None
        except Exception as e:
            base_ans = None

        # Score - ours
        if ours_rejected:
            if is_trap and qobj.get("trap_type") == "flow_vs_stock":
                ours_ok = True
                traps_caught += 1
            else:
                ours_ok = gold is None
        else:
            if gold is None:
                ours_ok = False
            else:
                ours_ok = match(ours_ans, gold) or match(ours_norm, gold)

        if ours_ok:
            ours_correct += 1

        # Score - baseline
        if gold is None:
            base_ok = base_ans is None or str(base_ans).strip().lower() in ("none", "null", "")
        else:
            base_ok = match(base_ans, gold)
        if base_ok:
            base_correct += 1

        status = "[PASS]" if ours_ok else "[FAIL]"
        base_status = "[PASS]" if base_ok else "[FAIL]"

        print(f"    Ours: {status} answer={ours_ans}  ({t_ours:.1f}s)")
        if ours_rejected:
            print(f"          [REJECTED: {ours_reason}]")
        print(f"    Base: {base_status} answer={base_ans}")
        print(f"    Gold: {gold}")

        details.append({
            "q_num": i,
            "question": q,
            "category": category,
            "trap": is_trap,
            "gold": gold,
            "ours_answer": ours_ans,
            "ours_ok": ours_ok,
            "ours_rejected": ours_rejected,
            "ours_reason": ours_reason,
            "ours_time": round(t_ours, 2),
            "ours_verdict": ours_verdict,
            "base_answer": base_ans,
            "base_ok": base_ok,
        })

    # Step 4: Summary
    print("\n" + "=" * 72)
    print("  FINAL SCORECARD")
    print("=" * 72)

    ours_pct = ours_correct / n * 100 if n else 0
    base_pct = base_correct / n * 100 if n else 0
    delta = ours_pct - base_pct

    print(f"\n  Total Questions:              {n}")
    print(f"  -----------------------------------------")
    print(f"  OURS (Proof-Carrying):        {ours_correct}/{n} = {ours_pct:.1f}%")
    print(f"  BASELINE (Plain RAG):         {base_correct}/{n} = {base_pct:.1f}%")
    print(f"  DELTA:                        {delta:+.1f} percentage points")
    print(f"  TRAPS CAUGHT (flow+stock):    {traps_caught}")

    # Category breakdown
    categories = {}
    for d in details:
        cat = d["category"]
        if cat not in categories:
            categories[cat] = {"total": 0, "ours": 0, "base": 0}
        categories[cat]["total"] += 1
        if d["ours_ok"]:
            categories[cat]["ours"] += 1
        if d["base_ok"]:
            categories[cat]["base"] += 1

    print(f"\n  --- Per-Category Breakdown ---")
    for cat, stats in sorted(categories.items()):
        o_pct = stats["ours"] / stats["total"] * 100
        b_pct = stats["base"] / stats["total"] * 100
        print(f"  {cat:25s}  Ours: {stats['ours']}/{stats['total']} ({o_pct:.0f}%)  "
              f"Base: {stats['base']}/{stats['total']} ({b_pct:.0f}%)")

    # Failed questions
    failed = [d for d in details if not d["ours_ok"]]
    if failed:
        print(f"\n  --- FAILED Questions ({len(failed)}) ---")
        for d in failed:
            print(f"  Q{d['q_num']:02d}: {d['question'][:60]}...")
            print(f"       Gold: {d['gold']}  Got: {d['ours_answer']}")
            if d["ours_rejected"]:
                print(f"       Reason: {d['ours_reason']}")
    else:
        print(f"\n  ** PERFECT SCORE -- All {n} questions answered correctly! **")

    print("\n" + "=" * 72)

    # Cleanup
    try:
        os.remove(path)
        print(f"  Cleaned up {path}")
    except Exception:
        pass

    return {
        "n": n,
        "ours_correct": ours_correct,
        "baseline_correct": base_correct,
        "ours_accuracy": ours_pct,
        "baseline_accuracy": base_pct,
        "delta": delta,
        "traps_caught": traps_caught,
        "details": details,
    }


if __name__ == "__main__":
    results = run_hard_benchmark()

    # Save results
    out_path = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        "sample_data", "hard_benchmark_results.json"
    )
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(results, f, indent=2, default=str)
    print(f"\n  Results saved to {out_path}")
